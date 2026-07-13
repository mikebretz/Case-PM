"""Extended estimating features — libraries, snapshots, Excel, award flows, integrations."""
from __future__ import annotations

import io
import json
import re
from datetime import datetime, timedelta

from sqlalchemy import inspect, text

from budget_persistence import get_budget_state, normalize_cost_code, save_budget_state
from estimate_persistence import (
    _line_extended,
    _parse_json,
    aggregate_lines_by_cost_code,
    apply_estimate_line_fields,
    estimate_line_to_dict,
    recalc_estimate_totals,
)

DEFAULT_RFP_TEMPLATE = {
    'subject': '{project_name} — RFP {package_number}: {package_title}',
    'body': (
        'You are invited to submit a bid for {package_title}.\n\n'
        'Spec section: {spec_section}\n'
        'Due date: {due_date}\n\n'
        'Please submit your quote or decline via the bid portal:\n'
        '{portal_url}\n\n'
        'Thank you.'
    ),
}

DEFAULT_ASSEMBLIES = [
    {
        'name': 'Interior partition wall (LF)',
        'trade': 'Drywall',
        'spec_section': '09 21 00',
        'unit': 'LF',
        'unit_cost': 0,
        'components': [
            {'description': 'Metal studs', 'unit': 'LF', 'factor': 1.0, 'cost_type': 'Material'},
            {'description': 'Drywall hang & finish', 'unit': 'LF', 'factor': 1.0, 'cost_type': 'Labor'},
        ],
    },
    {
        'name': 'Concrete slab (SF)',
        'trade': 'Concrete',
        'spec_section': '03 30 00',
        'unit': 'SF',
        'unit_cost': 0,
        'components': [
            {'description': 'Place & finish concrete', 'unit': 'SF', 'factor': 1.0, 'cost_type': 'Subcontract'},
            {'description': 'Vapor barrier', 'unit': 'SF', 'factor': 1.0, 'cost_type': 'Material'},
        ],
    },
]

RFP_NOTIFY_MODES = ('both', 'in_app', 'email', 'none')


def ensure_estimate_schema(engine, db):
    """Add columns/tables for extended estimating features."""
    insp = inspect(engine)
    tables = insp.get_table_names()

    def _add_col(table, name, col_type):
        if table not in tables:
            return
        cols = {c['name'] for c in insp.get_columns(table)}
        if name not in cols:
            db.session.execute(text(f'ALTER TABLE {table} ADD COLUMN {name} {col_type}'))
            db.session.commit()

    _add_col('estimate', 'settings_json', 'TEXT')
    _add_col('estimate', 'rom_amount', 'FLOAT DEFAULT 0')
    _add_col('bid_package', 'attachments_json', 'TEXT')
    _add_col('bid_package', 'email_template_json', 'TEXT')
    _add_col('estimate_line', 'line_kind', "VARCHAR(30) DEFAULT 'base'")
    _add_col('estimate_line', 'alternate_key', 'VARCHAR(40)')
    _add_col('estimate_line', 'assembly_id', 'INTEGER')
    _add_col('estimate_line', 'markup_id', 'INTEGER')
    _add_col('estimate_line', 'group_key', 'VARCHAR(80)')
    _add_col('estimate_line', 'meta_json', 'TEXT')
    _add_col('bid_invitation', 'qualification_json', 'TEXT')
    _add_col('bid_invitation', 'scope_gaps_json', 'TEXT')
    _add_col('bid_invitation', 'reminder_sent_at', 'DATETIME')

    db.create_all()


def get_estimate_settings(est):
    base = {
        'fee_breakdown_visible': True,
        'contingency_remaining': None,
        'rfp_email_template': DEFAULT_RFP_TEMPLATE,
        'budget_mapping_auto': True,
        'award_auto_commitment': False,
        'rfp_notify_mode': 'both',
        'ai_scope_enabled': True,
        'reminder_hours_before': 48,
    }
    merged = _parse_json(getattr(est, 'settings_json', None), {})
    base.update(merged)
    mode = str(base.get('rfp_notify_mode') or 'both').lower()
    if mode not in RFP_NOTIFY_MODES:
        mode = 'both'
    base['rfp_notify_mode'] = mode
    base['award_auto_commitment'] = bool(base.get('award_auto_commitment'))
    return base


def save_estimate_settings(est, patch):
    current = get_estimate_settings(est)
    current.update(patch or {})
    est.settings_json = json.dumps(current)
    est.updated_at = datetime.utcnow()


def line_to_dict_extended(line):
    d = estimate_line_to_dict(line)
    d.update({
        'line_kind': getattr(line, 'line_kind', None) or 'base',
        'alternate_key': getattr(line, 'alternate_key', None) or '',
        'assembly_id': getattr(line, 'assembly_id', None),
        'markup_id': getattr(line, 'markup_id', None),
        'group_key': getattr(line, 'group_key', None) or '',
        'meta': _parse_json(getattr(line, 'meta_json', None), {}),
    })
    return d


def apply_line_extended_fields(line, data):
    apply_estimate_line_fields(line, data)
    for field in ('line_kind', 'alternate_key', 'group_key'):
        if field in data:
            setattr(line, field, data[field] or '')
    if 'assembly_id' in data:
        line.assembly_id = data['assembly_id'] or None
    if 'markup_id' in data:
        line.markup_id = data['markup_id'] or None
    if 'meta' in data:
        line.meta_json = json.dumps(data['meta'] or {})


def included_alternate_keys(EstimateAlternate, estimate_id):
    rows = EstimateAlternate.query.filter_by(estimate_id=estimate_id, include_in_base=True).all()
    return {r.alt_key for r in rows if r.alt_key}


def line_counts_in_base(line, included_alt_keys):
    """Base total includes only alternates/allowances the owner marked include_in_base."""
    kind = (getattr(line, 'line_kind', None) or 'base').lower()
    alt_key = (getattr(line, 'alternate_key', None) or '').strip()
    if kind in ('alternate', 'allowance') or alt_key:
        return alt_key in included_alt_keys
    return True


def compute_fee_breakdown(est, lines, EstimateAlternate=None):
    estimate_id = getattr(est, 'id', None)
    included = included_alternate_keys(EstimateAlternate, estimate_id) if EstimateAlternate and estimate_id else set()
    direct = sum(
        float(l.extended_cost or 0)
        for l in lines
        if line_counts_in_base(l, included)
    )
    alternates_total = sum(
        float(l.extended_cost or 0)
        for l in lines
        if not line_counts_in_base(l, included)
    )
    cont_pct = float(est.contingency_pct or 0)
    oh_pct = float(est.overhead_pct or 0)
    profit_pct = float(est.profit_pct or 0)
    tax_pct = float(est.tax_pct or 0)
    contingency = direct * cont_pct / 100.0
    subtotal = direct + contingency
    overhead = subtotal * oh_pct / 100.0
    profit = subtotal * profit_pct / 100.0
    pre_tax = subtotal + overhead + profit
    tax = pre_tax * tax_pct / 100.0
    return {
        'direct_cost': round(direct, 2),
        'alternates_total': round(alternates_total, 2),
        'contingency': round(contingency, 2),
        'overhead': round(overhead, 2),
        'profit': round(profit, 2),
        'tax': round(tax, 2),
        'total': round(pre_tax + tax, 2),
    }


def assembly_to_dict(a):
    return {
        'id': a.id,
        'project_id': a.project_id,
        'name': a.name,
        'description': a.description or '',
        'trade': a.trade or '',
        'spec_section': a.spec_section or '',
        'unit': a.unit or 'EA',
        'unit_cost': float(a.unit_cost or 0),
        'components': _parse_json(a.components_json, []),
        'source': a.source or 'library',
    }


def seed_default_assemblies(EstimateAssembly, db, project_id=None):
    for row in DEFAULT_ASSEMBLIES:
        exists = EstimateAssembly.query.filter_by(name=row['name'], project_id=project_id).first()
        if exists:
            continue
        a = EstimateAssembly(
            project_id=project_id,
            name=row['name'],
            trade=row.get('trade'),
            spec_section=row.get('spec_section'),
            unit=row.get('unit', 'EA'),
            unit_cost=row.get('unit_cost', 0),
            components_json=json.dumps(row.get('components', [])),
            source='default',
        )
        db.session.add(a)
    db.session.commit()


def expand_assembly_to_lines(EstimateAssembly, assembly_id, qty, sort_start=0):
    """Return dict rows for worksheet from assembly components."""
    asm = EstimateAssembly.query.get(assembly_id)
    if not asm:
        return []
    components = _parse_json(asm.components_json, [])
    rows = []
    for i, comp in enumerate(components):
        factor = float(comp.get('factor') or 1)
        line_qty = float(qty or 1) * factor
        rows.append({
            'cost_code': asm.spec_section or '01-000',
            'spec_section': asm.spec_section or '',
            'description': f"{asm.name} — {comp.get('description', '')}",
            'cost_type': comp.get('cost_type') or 'Subcontract',
            'unit': comp.get('unit') or asm.unit or 'EA',
            'quantity': line_qty,
            'unit_cost': float(comp.get('unit_cost') or asm.unit_cost or 0),
            'source': 'assembly',
            'assembly_id': asm.id,
            'group_key': f'asm-{asm.id}',
            'sort_order': sort_start + i,
        })
    return rows


def snapshot_estimate(Estimate, EstimateLine, EstimateSnapshot, BidPackage, estimate_id, label, user_id, db):
    from estimate_persistence import estimate_to_dict, bid_package_to_dict
    est = Estimate.query.get_or_404(estimate_id)
    lines = EstimateLine.query.filter_by(estimate_id=estimate_id).all()
    packages = BidPackage.query.filter_by(estimate_id=estimate_id).all()
    payload = {
        'estimate': estimate_to_dict(est, [line_to_dict_extended(l) for l in lines]),
        'saved_at': datetime.utcnow().isoformat(),
    }
    snap = EstimateSnapshot(
        estimate_id=estimate_id,
        label=label or f"Rev {datetime.utcnow().strftime('%Y-%m-%d %H:%M')}",
        data_json=json.dumps(payload),
        created_by_id=user_id,
    )
    db.session.add(snap)
    db.session.commit()
    return snap


def list_snapshots(EstimateSnapshot, estimate_id):
    rows = EstimateSnapshot.query.filter_by(estimate_id=estimate_id).order_by(EstimateSnapshot.created_at.desc()).all()
    return [{'id': s.id, 'label': s.label, 'created_at': s.created_at.isoformat() if s.created_at else None} for s in rows]


def alternate_to_dict(a):
    return {
        'id': a.id,
        'alt_key': a.alt_key,
        'label': a.label,
        'include_in_base': bool(a.include_in_base),
        'amount': float(a.amount or 0),
        'notes': a.notes or '',
    }


def recalc_alternate_amounts(EstimateLine, EstimateAlternate, estimate_id):
    alts = EstimateAlternate.query.filter_by(estimate_id=estimate_id).all()
    for alt in alts:
        total = sum(
            float(l.extended_cost or 0)
            for l in EstimateLine.query.filter_by(estimate_id=estimate_id, alternate_key=alt.alt_key).all()
        )
        alt.amount = total
    return alts


def export_worksheet_excel(lines, alternates=None, fee_breakdown=None):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Worksheet'
    headers = ['Cost Code', 'Spec', 'Kind', 'Description', 'Cost Type', 'Qty', 'Unit', 'Unit Cost', 'Extended', 'Source', 'Group']
    ws.append(headers)
    for line in lines:
        ws.append([
            line.get('cost_code'), line.get('spec_section'), line.get('line_kind', 'base'),
            line.get('description'), line.get('cost_type'), line.get('quantity'),
            line.get('unit'), line.get('unit_cost'), line.get('extended_cost'),
            line.get('source'), line.get('group_key'),
        ])
    if alternates:
        ws2 = wb.create_sheet('Alternates')
        ws2.append(['Key', 'Label', 'Include in Base', 'Amount', 'Notes'])
        for a in alternates:
            ws2.append([a.get('alt_key'), a.get('label'), a.get('include_in_base'), a.get('amount'), a.get('notes')])
    if fee_breakdown:
        ws3 = wb.create_sheet('Fee Breakdown')
        for k, v in fee_breakdown.items():
            ws3.append([k, v])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def import_worksheet_excel(file_storage):
    from openpyxl import load_workbook
    wb = load_workbook(file_storage, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row or not any(row):
            continue
        rows.append({
            'cost_code': str(row[0] or ''),
            'spec_section': str(row[1] or ''),
            'line_kind': str(row[2] or 'base'),
            'description': str(row[3] or ''),
            'cost_type': str(row[4] or 'Subcontract'),
            'quantity': float(row[5] or 0),
            'unit': str(row[6] or 'EA'),
            'unit_cost': float(row[7] or 0),
            'source': str(row[9] or 'import'),
            'group_key': str(row[10] or ''),
            'sort_order': i,
        })
    return rows


def export_leveling_excel(matrix):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = 'Bid Leveling'
    vendors = set()
    for row in matrix:
        for b in row.get('bids', []):
            vendors.add(b.get('company_name'))
    vendor_list = sorted(vendors)
    ws.append(['Package', 'Spec'] + vendor_list + ['Low', 'Spread'])
    for row in matrix:
        pkg = row.get('package', {})
        bids = {b['company_name']: b['quote_amount'] for b in row.get('bids', [])}
        amounts = [bids.get(v, '') for v in vendor_list]
        quoted = [float(x) for x in amounts if x not in ('', None)]
        low = min(quoted) if quoted else ''
        spread = (max(quoted) - min(quoted)) if len(quoted) > 1 else ''
        ws.append([pkg.get('number'), pkg.get('spec_section')] + amounts + [low, spread])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def check_vendor_qualification(Company, COI, company_id, company_name):
    """Return qualification flags for vendor invite."""
    company = None
    if company_id:
        try:
            company = Company.query.get(int(company_id))
        except (TypeError, ValueError):
            company = None
    if not company and company_name:
        company = Company.query.filter_by(name=company_name).first()
    result = {'coi_valid': False, 'licensed': False, 'qualified': False, 'notes': ''}
    if not company:
        result['notes'] = 'Company not found in directory'
        return result
    result['licensed'] = bool(company.license_number)
    coi = COI.query.filter_by(company_id=company.id).order_by(COI.expiration_date.desc()).first() if COI else None
    if coi and coi.expiration_date and coi.expiration_date >= datetime.utcnow().date():
        result['coi_valid'] = True
    details = _parse_json(company.details_json, {})
    result['qualified'] = bool(details.get('prequalified') or details.get('qualified'))
    if not result['qualified']:
        result['qualified'] = result['coi_valid'] and result['licensed']
    return result


def render_rfp_email(template, project, package, portal_url):
    subj = template.get('subject', DEFAULT_RFP_TEMPLATE['subject'])
    body = template.get('body', DEFAULT_RFP_TEMPLATE['body'])
    repl = {
        '{project_name}': getattr(project, 'name', '') or '',
        '{package_number}': package.number or '',
        '{package_title}': package.title or '',
        '{spec_section}': package.spec_section or '',
        '{due_date}': package.due_date.isoformat() if package.due_date else 'TBD',
        '{portal_url}': portal_url,
    }
    for k, v in repl.items():
        subj = subj.replace(k, v)
        body = body.replace(k, v)
    return subj, body


def send_bid_reminders(BidPackage, BidInvitation, Project, User, db, hours_before=48, Estimate=None):
    """Notify vendors approaching due date and staff on overdue quotes."""
    from estimate_persistence import notify_bid_invitations
    from estimate_features import get_estimate_settings
    now = datetime.utcnow()
    sent = 0
    packages = BidPackage.query.filter(BidPackage.status == 'Open').all()
    for pkg in packages:
        if not pkg.due_date:
            continue
        due_dt = datetime.combine(pkg.due_date, datetime.min.time())
        hours_left = (due_dt - now).total_seconds() / 3600.0
        invitations = BidInvitation.query.filter_by(bid_package_id=pkg.id, status='Sent').all()
        if 0 < hours_left <= hours_before:
            pending = [i for i in invitations if not i.reminder_sent_at]
            if pending:
                est = Estimate.query.get(pkg.estimate_id) if Estimate and getattr(pkg, 'estimate_id', None) else None
                notify_mode = get_estimate_settings(est).get('rfp_notify_mode') if est else 'both'
                notify_bid_invitations(
                    pkg.project_id, pending, pkg, User,
                    title=f'Reminder: {pkg.number} due {pkg.due_date}',
                    notify_mode=notify_mode,
                    estimate=est,
                    Project=Project,
                )
                for inv in pending:
                    inv.reminder_sent_at = now
                sent += len(pending)
        elif hours_left < 0:
            quoted = BidInvitation.query.filter_by(bid_package_id=pkg.id, status='Quoted').count()
            if quoted < len(invitations):
                try:
                    from email_notifications import notify_role_workflow
                    notify_role_workflow(
                        User, 'Project Manager',
                        title=f'Overdue bids: {pkg.number}',
                        description=f'{pkg.title} was due {pkg.due_date}. {quoted}/{len(invitations)} quotes received.',
                        action_url=f'/estimating?project_id={pkg.project_id}&tab=leveling',
                        project_id=pkg.project_id,
                        module='Estimating',
                    )
                except Exception:
                    pass
    db.session.commit()
    return sent


def lookup_historical_cost(EstimateCostHistory, cost_code=None, trade=None, unit=None, project_id=None):
    q = EstimateCostHistory.query
    if project_id:
        q = q.filter((EstimateCostHistory.project_id == project_id) | (EstimateCostHistory.project_id.is_(None)))
    rows = q.order_by(EstimateCostHistory.recorded_at.desc()).limit(200).all()
    if cost_code:
        target = normalize_cost_code(cost_code)
        rows = [r for r in rows if normalize_cost_code(r.cost_code) == target]
    if trade:
        t = trade.lower()
        rows = [r for r in rows if t in (r.trade or '').lower()]
    if unit:
        rows = [r for r in rows if (r.unit or '').lower() == unit.lower()]
    return [{
        'cost_code': r.cost_code,
        'trade': r.trade,
        'unit': r.unit,
        'unit_cost': float(r.unit_cost or 0),
        'description': r.description,
        'source_project_name': r.source_project_name,
        'recorded_at': r.recorded_at.isoformat() if r.recorded_at else None,
    } for r in rows[:20]]


def ai_extract_scope_from_spec(text_content, spec_section=None):
    """Heuristic scope extraction stub — returns suggested line items."""
    if not text_content:
        return []
    lines = []
    section = spec_section or '01 00 00'
    for match in re.finditer(r'(\d+(?:\.\d+)?)\s*(SF|SY|LF|EA|CY|TON)\b', text_content, re.I):
        qty, unit = match.groups()
        lines.append({
            'spec_section': section,
            'description': f'AI suggested — {qty} {unit.upper()} from spec excerpt',
            'quantity': float(qty),
            'unit': unit.upper(),
            'cost_code': section,
            'source': 'ai_scope',
            'confidence': 0.6,
        })
    if not lines and len(text_content) > 40:
        lines.append({
            'spec_section': section,
            'description': f'AI scope summary — {text_content[:120].strip()}…',
            'quantity': 1,
            'unit': 'LS',
            'cost_code': section,
            'source': 'ai_scope',
            'confidence': 0.4,
        })
    return lines[:15]


def apply_budget_mappings(EstimateBudgetMapping, project_id, rollups):
    mappings = EstimateBudgetMapping.query.filter_by(project_id=project_id).all()
    if not mappings:
        return rollups
    map_by_spec = {m.spec_section.replace(' ', '').upper(): m for m in mappings}
    out = []
    for row in rollups:
        spec = (row.get('spec_section') or row.get('cost_code') or '').replace(' ', '').upper()
        m = map_by_spec.get(spec)
        if m:
            row = dict(row)
            row['cost_code'] = m.cost_code
            row['cost_type'] = m.cost_type
        out.append(row)
    return out


def award_to_commitment_draft(
    BidPackage, BidInvitation, Commitment, CommitmentAllocation, db,
    package_id, generate_number_fn, user_id,
):
    pkg = BidPackage.query.get_or_404(package_id)
    inv = BidInvitation.query.get(pkg.awarded_invitation_id) if pkg.awarded_invitation_id else None
    if not inv:
        raise ValueError('No awarded vendor for this package')
    c = Commitment(
        project_id=pkg.project_id,
        number=generate_number_fn('SC', Commitment, doc_type='commitment_sc', project_id=pkg.project_id),
        title=pkg.title,
        description=pkg.scope_notes or pkg.description or pkg.title,
        commitment_type='Subcontract',
        status='Draft',
        company_name=inv.company_name,
        company_id=inv.company_id,
        original_amount=float(inv.quote_amount or 0),
        current_amount=float(inv.quote_amount or 0),
        ball_in_court_role='Creator',
        created_by_id=user_id,
    )
    db.session.add(c)
    db.session.flush()
    alloc = CommitmentAllocation(
        commitment_id=c.id,
        cost_code=pkg.spec_section or '01-000',
        amount=float(inv.quote_amount or 0),
        description=pkg.title,
    )
    db.session.add(alloc)
    db.session.commit()
    return c


def sync_estimate_to_forecast(Estimate, BudgetProjectState, db, estimate_id, user_id=None):
    """Push ROM / estimate total into budget metadata for forecast chain."""
    est = Estimate.query.get_or_404(estimate_id)
    _, state = get_budget_state(BudgetProjectState, est.project_id)
    state['estimate_rom'] = float(est.total_amount or est.rom_amount or 0)
    state['estimate_id'] = est.id
    state['estimate_number'] = est.number
    state['estimate_synced_at'] = datetime.utcnow().isoformat()
    save_budget_state(BudgetProjectState, db, est.project_id, state, user_id)
    est.rom_amount = float(est.total_amount or 0)
    db.session.commit()
    return state.get('estimate_rom')


def apply_contingency_drawdown(BudgetProjectState, db, project_id, amount, user_id=None, note=''):
    """Release contingency from budget state."""
    _, state = get_budget_state(BudgetProjectState, project_id)
    lines = list(state.get('budgetLines') or [])
    cont_line = next((l for l in lines if 'contingency' in (l.get('description') or '').lower()), None)
    if not cont_line:
        cont_line = {
            'id': int(datetime.utcnow().timestamp() * 1000),
            'cost_code': '01-0000',
            'description': 'Contingency',
            'cost_type': 'Other',
            'original_budget': 0,
            'approved_changes': 0,
            'pending': 0,
            'actual': 0,
            'notes': '',
            'syncStatus': 'Not Synced',
            'percent_complete': 0,
        }
        lines.append(cont_line)
    amt = float(amount or 0)
    cont_line['original_budget'] = max(0, float(cont_line.get('original_budget') or 0) - amt)
    cont_line['notes'] = (cont_line.get('notes') or '') + f"\nDrawdown {amt}: {note}"
    state['budgetLines'] = lines
    audit = list(state.get('budgetAuditLog') or [])
    audit.append({'timestamp': datetime.utcnow().isoformat(), 'action': 'CONTINGENCY_DRAWDOWN', 'amount': amt, 'note': note})
    state['budgetAuditLog'] = audit
    save_budget_state(BudgetProjectState, db, project_id, state, user_id)
    return cont_line


def filter_group_lines(lines, filters=None):
    filters = filters or {}
    q = (filters.get('q') or '').strip().lower()
    kind = filters.get('kind')
    source = filters.get('source')
    division = filters.get('division')
    unpriced = filters.get('unpriced')
    group = filters.get('group')
    out = []
    for line in lines:
        if q and not any(q in str(line.get(k) or '').lower() for k in ('cost_code', 'spec_section', 'description', 'group_key')):
            continue
        if kind and line.get('line_kind') != kind:
            continue
        if source and line.get('source') != source:
            continue
        if division and line.get('division') != division:
            continue
        if unpriced and float(line.get('unit_cost') or 0) > 0:
            continue
        if group and line.get('group_key') != group:
            continue
        out.append(line)
    groups = {}
    for line in out:
        gk = line.get('group_key') or line.get('division') or 'Ungrouped'
        groups.setdefault(gk, {'key': gk, 'lines': [], 'subtotal': 0})
        groups[gk]['lines'].append(line)
        groups[gk]['subtotal'] += float(line.get('extended_cost') or 0)
    return {'lines': out, 'groups': list(groups.values()), 'count': len(out)}


def build_dashboard_estimating_tile(Estimate, BidPackage, BidInvitation, project_id):
    if not project_id:
        return {}
    estimates = Estimate.query.filter_by(project_id=int(project_id)).all()
    open_packages = BidPackage.query.filter_by(project_id=int(project_id)).filter(BidPackage.status == 'Open').all()
    due_soon = 0
    today = datetime.utcnow().date()
    for p in open_packages:
        if p.due_date and today <= p.due_date <= today + timedelta(days=7):
            due_soon += 1
    invites = BidInvitation.query.join(BidPackage).filter(BidPackage.project_id == int(project_id)).all()
    quotes = sum(1 for i in invites if i.status in ('Quoted', 'Awarded'))
    active = next((e for e in estimates if e.status not in ('Archived', 'Lost')), estimates[0] if estimates else None)
    return {
        'estimate_count': len(estimates),
        'active_estimate': active.number if active else None,
        'active_total': float(active.total_amount or 0) if active else 0,
        'open_packages': len(open_packages),
        'quotes_received': quotes,
        'due_this_week': due_soon,
    }


def import_takeoff_with_markup(EstimateLine, estimate_id, takeoff_items, db, default_cost_code='01-000', assembly_id=None):
    """Enhanced takeoff import storing markup_id for bi-directional link."""
    imported = []
    existing = {l.markup_id for l in EstimateLine.query.filter_by(estimate_id=estimate_id).all() if l.markup_id}
    sort_base = EstimateLine.query.filter_by(estimate_id=estimate_id).count()
    for i, item in enumerate(takeoff_items):
        mid = item.get('markup_id')
        if mid and mid in existing:
            continue
        qty = float(item.get('quantity') or 0)
        line = EstimateLine(
            estimate_id=estimate_id,
            sort_order=sort_base + i,
            cost_code=default_cost_code,
            description=(item.get('description') or f"Takeoff {item.get('sheet_number', '')}")[:500],
            cost_type='Subcontract',
            unit=item.get('unit') or 'ft',
            quantity=qty,
            unit_cost=0,
            extended_cost=0,
            source='takeoff',
            source_ref=f'takeoff:{mid}',
            markup_id=mid,
            notes=f"Sheet {item.get('sheet_number', '')}",
        )
        if assembly_id:
            line.assembly_id = assembly_id
            line.group_key = f'asm-{assembly_id}'
        db.session.add(line)
        imported.append(line)
    return imported
