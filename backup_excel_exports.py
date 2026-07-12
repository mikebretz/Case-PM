"""Generate human-readable Excel exports bundled inside program backup zips.

These spreadsheets are for data portability (import into other systems).
They are NOT used when restoring a backup — restore still uses case_pm.db only.
"""
from __future__ import annotations

import json
import os
import re
from datetime import date, datetime
from typing import Any

EXCEL_ROOT = 'excel_exports'


def _cell(value: Any) -> Any:
    if value is None:
        return ''
    if isinstance(value, bool):
        return 'Yes' if value else 'No'
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (dict, list)):
        try:
            return json.dumps(value, ensure_ascii=False)[:32000]
        except (TypeError, ValueError):
            return str(value)[:32000]
    return value


def _slug(text: str, fallback: str = 'item') -> str:
    raw = re.sub(r'[^a-zA-Z0-9_-]+', '_', (text or '').strip())[:60].strip('_')
    return raw or fallback


def _project_slug(project) -> str:
    parts = [str(getattr(project, 'number', '') or project.id), getattr(project, 'name', '') or '']
    return _slug('_'.join(p for p in parts if p), f'project_{project.id}')


def _write_workbook(path: str, sheets: list[tuple[str, list[str], list[dict]]]) -> None:
    from openpyxl import Workbook

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, headers, rows in sheets:
        title = (sheet_name or 'Sheet')[:31]
        ws = wb.create_sheet(title=title)
        ws.append(headers)
        for row in rows:
            ws.append([_cell(row.get(h)) for h in headers])
    if not wb.sheetnames:
        ws = wb.create_sheet('Empty')
        ws.append(['No data'])
    wb.save(path)


def _rows_from_dicts(items: list[dict], columns: list[str] | None = None) -> tuple[list[str], list[dict]]:
    if not items:
        cols = columns or ['Note']
        return cols, [{'Note': 'No records'}]
    if columns:
        headers = columns
    else:
        headers = []
        for item in items:
            for key in item.keys():
                if key not in headers:
                    headers.append(key)
    return headers, items


def _parse_json(raw, default=None):
    if default is None:
        default = {}
    if not raw:
        return default
    try:
        return json.loads(raw)
    except (TypeError, json.JSONDecodeError):
        return default


def build_excel_exports_to_dir(dest_root: str, models: dict, progress_cb=None, db=None) -> dict:
    """Write excel_exports tree under dest_root. Returns summary stats."""
    from budget_persistence import get_budget_state
    from pay_app_persistence import get_pay_app_state

    if db is not None:
        try:
            from companies_persistence import ensure_company_schema
            ensure_company_schema(db)
        except Exception:
            pass

    def step(pct, msg):
        if progress_cb:
            progress_cb(pct, msg)

    Project = models['Project']
    projects = Project.query.order_by(Project.number, Project.name).all()
    stats = {'files': 0, 'projects': len(projects)}

    readme = os.path.join(dest_root, 'README.txt')
    os.makedirs(dest_root, exist_ok=True)
    with open(readme, 'w', encoding='utf-8') as fh:
        fh.write(
            'Case PM — Excel data exports (portability copies)\n'
            '================================================\n\n'
            'These spreadsheets mirror the tabular data from each module so you can\n'
            'review or import into another system. They are included inside every\n'
            'program backup zip for convenience.\n\n'
            'IMPORTANT: Restoring a backup uses case_pm.db and uploads/ only.\n'
            'These Excel files are NOT loaded back into Case PM automatically.\n\n'
            'Folder layout:\n'
            '  program/     — company-wide lists (projects, companies, users)\n'
            '  by_project/  — one subfolder per job with module spreadsheets\n'
        )

    step(58, 'Building Excel exports — program lists…')
    try:
        _export_program_lists(dest_root, models)
        stats['files'] += 3
    except Exception as exc:
        stats['program_lists_error'] = str(exc)

    for idx, project in enumerate(projects, start=1):
        slug = _project_slug(project)
        pct = 58 + int((idx / max(len(projects), 1)) * 28)
        step(pct, f'Excel exports ({idx}/{len(projects)}): {slug}…')
        proj_dir = os.path.join(dest_root, 'by_project', slug)
        try:
            count = _export_project_modules(
                proj_dir, project, models,
                get_budget_state=get_budget_state,
                get_pay_app_state=get_pay_app_state,
            )
            stats['files'] += count
        except Exception as exc:
            stats.setdefault('project_errors', []).append({'project': slug, 'error': str(exc)})

    step(87, f'Excel exports complete ({stats["files"]} files)')
    return stats


def _export_program_lists(dest_root: str, models: dict) -> None:
    program_dir = os.path.join(dest_root, 'program')
    Project = models['Project']
    Company = models['Company']
    User = models['User']

    projects = Project.query.order_by(Project.number).all()
    proj_rows = [{
        'Number': p.number,
        'Name': p.name,
        'Client': p.client,
        'Status': p.status,
        'City': p.city,
        'State': p.state,
        'Start Date': p.start_date,
        'End Date': p.end_date,
        'Contract Value': p.contract_value,
        'Percent Complete': p.percent_complete,
        'Project Manager': p.project_manager,
        'Sage Job Number': p.sage_job_number,
    } for p in projects]
    headers, rows = _rows_from_dicts(proj_rows)
    _write_workbook(os.path.join(program_dir, 'Projects.xlsx'), [('Projects', headers, rows)])

    companies = Company.query.order_by(Company.name).all()
    from companies_persistence import serialize_company
    comp_rows = []
    for c in companies:
        data = serialize_company(c)
        comp_rows.append({
            'Name': data.get('company_name') or data.get('name'),
            'Type': data.get('company_type') or data.get('type'),
            'Trade': data.get('trade'),
            'Email': data.get('primary_email') or data.get('email'),
            'Phone': data.get('primary_phone') or data.get('phone'),
            'License': data.get('license_number'),
            'Tax ID': data.get('tax_id'),
            'Status': data.get('status'),
            'Sage Number': data.get('sage_number'),
        })
    headers, rows = _rows_from_dicts(comp_rows)
    _write_workbook(os.path.join(program_dir, 'Companies.xlsx'), [('Companies', headers, rows)])

    users = User.query.order_by(User.email).all()
    user_rows = [{
        'First Name': u.first_name,
        'Last Name': u.last_name,
        'Email': u.email,
        'Role': u.role,
        'Status': u.status,
        'Job Title': getattr(u, 'job_title', None),
        'Phone': getattr(u, 'phone', None),
    } for u in users]
    headers, rows = _rows_from_dicts(user_rows)
    _write_workbook(os.path.join(program_dir, 'Users.xlsx'), [('Users', headers, rows)])


def _export_project_modules(proj_dir, project, models, *, get_budget_state, get_pay_app_state) -> int:
    from co_persistence import co_to_dict
    from rfi_persistence import rfi_to_dict, get_linked_records
    from punch_persistence import serialize_item
    from daily_log_persistence import serialize_log
    from weekly_report_persistence import serialize_report
    from safety_persistence import serialize_report as serialize_safety_report, serialize_cert
    from meeting_minutes_persistence import serialize_meeting
    from photo_persistence import serialize_photo
    from deliveries_persistence import serialize_delivery
    from permits_inspections_persistence import serialize_item as serialize_inspection

    pid = project.id
    files_written = 0
    os.makedirs(proj_dir, exist_ok=True)

    RFI = models['RFI']
    ChangeOrder = models['ChangeOrder']
    PotentialChangeOrder = models['PotentialChangeOrder']
    Commitment = models['Commitment']
    Submittal = models['Submittal']
    PunchItem = models['PunchItem']
    DailyLog = models['DailyLog']
    ManpowerEntry = models['ManpowerEntry']
    EquipmentEntry = models['EquipmentEntry']
    WeeklyReport = models['WeeklyReport']
    SafetyReport = models['SafetyReport']
    SafetyCertification = models['SafetyCertification']
    ScheduleData = models['ScheduleData']
    ScheduleTask = models['ScheduleTask']
    Delivery = models['Delivery']
    PermitInspectionItem = models['PermitInspectionItem']
    MeetingMinute = models['MeetingMinute']
    MeetingActionItem = models['MeetingActionItem']
    Photo = models['Photo']
    Document = models['Document']
    Drawing = models['Drawing']
    BudgetProjectState = models['BudgetProjectState']
    PayAppProjectState = models['PayAppProjectState']
    User = models.get('User')

    def save(name: str, sheets: list[tuple[str, list[str], list[dict]]]) -> None:
        nonlocal files_written
        _write_workbook(os.path.join(proj_dir, name), sheets)
        files_written += 1

    # RFIs
    rfi_rows = []
    for rfi in RFI.query.filter_by(project_id=pid).order_by(RFI.number).all():
        linked_cos, linked_pcos = get_linked_records(rfi.id, ChangeOrder, PotentialChangeOrder)
        d = rfi_to_dict(rfi, linked_cos, linked_pcos)
        rfi_rows.append({
            'Number': d.get('number'), 'Subject': d.get('subject'), 'Status': d.get('status'),
            'Priority': d.get('priority'), 'Ball In Court': d.get('ball_in_court_role'),
            'Due Date': d.get('due_date'), 'Drawing': d.get('drawing_reference'),
            'Spec': d.get('spec_reference'), 'Company': d.get('received_from_company'),
            'Question': d.get('question'), 'Answer': d.get('official_answer'),
            'Cost Impact': d.get('cost_impact_amount'), 'Schedule Days': d.get('schedule_impact_days'),
        })
    h, rows = _rows_from_dicts(rfi_rows)
    save('RFIs.xlsx', [('RFI Log', h, rows)])

    # Change orders + PCOs
    co_rows = []
    for co in ChangeOrder.query.filter_by(project_id=pid).order_by(ChangeOrder.number).all():
        d = co_to_dict(co)
        co_rows.append({
            'Number': d.get('number'), 'Date': d.get('date'), 'Title': d.get('title'),
            'Company': d.get('company_name'), 'Amount': d.get('amount'), 'Status': d.get('status'),
            'Ball In Court': d.get('ball_in_court_role'), 'Schedule Days': d.get('schedule_impact_days'),
            'SOV Synced': 'Yes' if d.get('sov_synced_at') else 'No',
        })
    h, rows = _rows_from_dicts(co_rows)
    save('Change_Orders.xlsx', [('Change Orders', h, rows)])

    pco_rows = []
    for pco in PotentialChangeOrder.query.filter_by(project_id=pid).order_by(PotentialChangeOrder.number).all():
        pco_rows.append({
            'Number': pco.number, 'Title': pco.title, 'Status': pco.status,
            'Estimated Amount': pco.estimated_amount, 'Reason': pco.reason,
            'Priority': pco.priority, 'Company': pco.company_name,
            'Schedule Days': pco.schedule_impact_days,
        })
    h, rows = _rows_from_dicts(pco_rows)
    save('PCOs.xlsx', [('PCOs', h, rows)])

    # Commitments
    com_rows = []
    for c in Commitment.query.filter_by(project_id=pid).order_by(Commitment.number).all():
        com_rows.append({
            'Number': c.number, 'Type': c.commitment_type, 'Title': c.title,
            'Vendor': c.company_name, 'AIA Form': c.aia_form,
            'Original': c.original_amount, 'Changes': c.approved_changes,
            'Current': c.current_amount, 'Status': c.status,
            'Signature': c.signature_status, 'Sage': c.sage_sync_status,
        })
    h, rows = _rows_from_dicts(com_rows)
    save('Commitments.xlsx', [('Commitments', h, rows)])

    # Budget
    _, budget_state = get_budget_state(BudgetProjectState, pid)
    budget_lines = budget_state.get('budgetLines') or []
    budget_rows = []
    for i, line in enumerate(budget_lines, start=1):
        if not isinstance(line, dict):
            continue
        orig = float(line.get('original_budget') or 0)
        approved = float(line.get('approved_changes') or 0)
        pending = float(line.get('pending') or 0)
        revised = orig + approved + pending
        actual = float(line.get('actual') or 0)
        budget_rows.append({
            '#': i,
            'Cost Code': line.get('cost_code'),
            'Description': line.get('description'),
            'Cost Type': line.get('cost_type'),
            'Original Budget': orig,
            'Approved COs': approved,
            'Pending Changes': pending,
            'Committed': line.get('committed') or 0,
            'Revised Budget': revised,
            'Actual Cost': actual,
            'Variance': revised - actual,
            '% Complete': line.get('percent_complete') or 0,
            'Notes': line.get('notes') or '',
        })
    h, rows = _rows_from_dicts(budget_rows)
    save('Budget.xlsx', [('Budget', h, rows)])

    # Pay applications — contractor SOV
    _, pay_state = get_pay_app_state(PayAppProjectState, pid)
    sov_rows = []
    for i, line in enumerate(pay_state.get('contractorSOV') or [], start=1):
        if not isinstance(line, dict):
            continue
        sov_rows.append({
            '#': i,
            'Cost Code': line.get('costCode') or line.get('cost_code'),
            'Description': line.get('description'),
            'Scheduled Value': line.get('scheduledValue') or line.get('scheduled_value'),
            'Work Completed From Previous': line.get('workCompletedFromPreviousApplications'),
            'Work Completed This Period': line.get('workCompletedThisPeriod'),
            'Materials Stored': line.get('materialsPresentlyStored'),
            'Total Completed Stored': line.get('totalCompletedAndStored'),
            '% Complete': line.get('percentComplete'),
            'Balance To Finish': line.get('balanceToFinish'),
            'Retainage': line.get('retainage'),
        })
    h, rows = _rows_from_dicts(sov_rows)
    save('Pay_Applications_SOV.xlsx', [('Contractor SOV', h, rows)])

    # Submittals
    sub_rows = []
    for s in Submittal.query.filter_by(project_id=pid).order_by(Submittal.number).all():
        sub_rows.append({
            'Number': s.number, 'Description': s.description, 'Spec Section': s.spec_section,
            'Status': s.status, 'Priority': s.priority, 'Submitted By': s.submitted_by,
            'Date': s.date, 'Due Date': s.due_date, 'Ball In Court': s.ball_in_court,
        })
    h, rows = _rows_from_dicts(sub_rows)
    save('Submittals.xlsx', [('Submittals', h, rows)])

    # Punch list
    punch_rows = []
    for item in PunchItem.query.filter_by(project_id=pid).order_by(PunchItem.number).all():
        d = serialize_item(item, User=User, summary=True)
        punch_rows.append({
            'Number': d.get('number'), 'Description': d.get('description'),
            'Location': d.get('location'), 'Trade': d.get('trade'),
            'Category': d.get('category'), 'Priority': d.get('priority'),
            'Status': d.get('status'), 'Due Date': d.get('due_date'),
            'Assigned To': d.get('assigned_to'), 'Company': d.get('assigned_company'),
        })
    h, rows = _rows_from_dicts(punch_rows)
    save('Punch_List.xlsx', [('Punch List', h, rows)])

    # Daily log
    log_rows = []
    for log in DailyLog.query.filter_by(project_id=pid).order_by(DailyLog.date.desc()).all():
        d = serialize_log(log, ManpowerEntry, EquipmentEntry, User=User, summary=True)
        log_rows.append({
            'Date': d.get('date'), 'Weather': d.get('weather'),
            'Work Performed': d.get('work_performed'), 'Notes': d.get('notes'),
            'Status': d.get('status'), 'Author': d.get('author'),
            'Workers': d.get('total_workers'), 'Hours': d.get('total_hours'),
        })
    h, rows = _rows_from_dicts(log_rows)
    save('Daily_Log.xlsx', [('Daily Log', h, rows)])

    # Weekly reports
    wr_rows = []
    for r in WeeklyReport.query.filter_by(project_id=pid).order_by(WeeklyReport.week_ending.desc()).all():
        d = serialize_report(r, User=User, summary=True)
        wr_rows.append({
            'Week Ending': d.get('week_ending'), 'Period Type': d.get('period_type'),
            'Status': d.get('status'), 'Work Performed': d.get('work_performed'),
            'Safety Notes': d.get('safety_notes'),
        })
    h, rows = _rows_from_dicts(wr_rows)
    save('Weekly_Reports.xlsx', [('Weekly Reports', h, rows)])

    # Safety
    safety_rows = []
    for r in SafetyReport.query.filter_by(project_id=pid).order_by(SafetyReport.created_at.desc()).all():
        d = serialize_safety_report(r, User=User, summary=True)
        safety_rows.append({
            'Number': d.get('number'), 'Type': d.get('type'),
            'Description': d.get('description'), 'Location': d.get('location'),
            'Severity': d.get('severity'), 'Status': d.get('status'),
            'Report Date': d.get('report_date'),
        })
    h, rows = _rows_from_dicts(safety_rows)
    save('Safety_Reports.xlsx', [('Safety Reports', h, rows)])

    cert_rows = []
    for c in SafetyCertification.query.filter_by(project_id=pid).all():
        d = serialize_cert(c, summary=True)
        cert_rows.append({
            'Person': d.get('person_name'), 'Company': d.get('company'),
            'Cert Type': d.get('cert_type'), 'Expiration': d.get('expiration_date'),
            'Card Number': d.get('card_number'),
        })
    h, rows = _rows_from_dicts(cert_rows)
    save('Safety_Certifications.xlsx', [('Certifications', h, rows)])

    # Schedule
    sched_rows = []
    sched_data = ScheduleData.query.filter_by(project_id=pid).first()
    if sched_data and sched_data.payload:
        payload = _parse_json(sched_data.payload, {})
        tasks = payload.get('data') or payload.get('tasks') or []
        if isinstance(tasks, list):
            for t in tasks:
                if not isinstance(t, dict):
                    continue
                sched_rows.append({
                    'ID': t.get('id'), 'Number': t.get('number') or t.get('text'),
                    'Description': t.get('text') or t.get('description'),
                    'Start': t.get('start_date') or t.get('start'),
                    'End': t.get('end_date') or t.get('end'),
                    'Duration': t.get('duration'),
                    '% Complete': t.get('progress') or t.get('percent_complete'),
                    'Status': t.get('status'),
                })
    if not sched_rows:
        for t in ScheduleTask.query.filter_by(project_id=pid).order_by(ScheduleTask.number).all():
            sched_rows.append({
                'Number': t.number, 'Description': t.description, 'Phase': t.phase,
                'Start': t.start_date, 'End': t.end_date, 'Duration Days': t.duration_days,
                '% Complete': t.percent_complete, 'Status': t.status,
                'Assigned To': t.assigned_to,
            })
    h, rows = _rows_from_dicts(sched_rows)
    save('Schedule.xlsx', [('Schedule', h, rows)])

    # Deliveries
    del_rows = []
    for d in Delivery.query.filter_by(project_id=pid).order_by(Delivery.delivery_date).all():
        row = serialize_delivery(d, User=User)
        del_rows.append({
            'Number': row.get('delivery_number'), 'Supplier': row.get('supplier'),
            'Description': row.get('description'), 'Date': row.get('delivery_date'),
            'Time Window': row.get('time_window'), 'Status': row.get('status'),
            'Location': row.get('location'), 'PO Number': row.get('po_number'),
        })
    h, rows = _rows_from_dicts(del_rows)
    save('Deliveries.xlsx', [('Deliveries', h, rows)])

    # Permits & inspections
    insp_rows = []
    for item in PermitInspectionItem.query.filter_by(project_id=pid).order_by(PermitInspectionItem.scheduled_date).all():
        row = serialize_inspection(item, User=User)
        insp_rows.append({
            'Number': row.get('item_number'), 'Kind': row.get('record_kind'),
            'Title': row.get('title'), 'Trade': row.get('trade'),
            'Phase': row.get('inspection_phase'), 'Scheduled': row.get('scheduled_date'),
            'Status': row.get('status'), 'Permit Number': row.get('permit_number'),
            'Jurisdiction': row.get('jurisdiction_name'),
        })
    h, rows = _rows_from_dicts(insp_rows)
    save('Permits_Inspections.xlsx', [('Permits Inspections', h, rows)])

    # Meeting minutes
    meet_rows = []
    for m in MeetingMinute.query.filter_by(project_id=pid).order_by(MeetingMinute.meeting_date.desc()).all():
        row = serialize_meeting(m, include_actions=False, ActionItem=MeetingActionItem)
        meet_rows.append({
            'Number': row.get('meeting_number'), 'Date': row.get('meeting_date'),
            'Type': row.get('meeting_type'), 'Subject': row.get('subject'),
            'Status': row.get('status'), 'Location': row.get('location'),
            'Organizer': row.get('organizer'),
        })
    h, rows = _rows_from_dicts(meet_rows)
    save('Meeting_Minutes.xlsx', [('Meeting Minutes', h, rows)])

    # Photos (metadata)
    photo_rows = []
    for p in Photo.query.filter_by(project_id=pid).order_by(Photo.taken_date.desc()).all():
        row = serialize_photo(p, user=User)
        photo_rows.append({
            'Filename': row.get('filename'), 'Caption': row.get('caption'),
            'Location': row.get('location'), 'Category': row.get('category'),
            'Taken Date': row.get('taken_date'),
        })
    h, rows = _rows_from_dicts(photo_rows)
    save('Photos.xlsx', [('Photos', h, rows)])

    # Documents index
    doc_rows = []
    for doc in Document.query.filter_by(project_id=pid).order_by(Document.name).all():
        doc_rows.append({
            'Name': doc.name, 'Type': doc.document_type,
            'Filename': doc.filename, 'Folder ID': doc.folder_id,
            'Size Bytes': doc.file_size, 'Created': doc.created_at,
            'Updated': doc.updated_at,
        })
    h, rows = _rows_from_dicts(doc_rows)
    save('Documents_Index.xlsx', [('Documents', h, rows)])

    # Drawings index
    draw_rows = []
    for dr in Drawing.query.filter_by(project_id=pid).order_by(Drawing.sheet_number).all():
        draw_rows.append({
            'Sheet Number': dr.sheet_number, 'Title': dr.title,
            'Discipline': dr.discipline, 'Status': dr.status,
            'Section': dr.section_prefix,
        })
    h, rows = _rows_from_dicts(draw_rows)
    save('Drawings_Index.xlsx', [('Drawings', h, rows)])

    return files_written


def get_backup_export_models() -> dict:
    """Lazy import of SQLAlchemy models — call inside Flask app context."""
    from app import (
        AuditLog,
        BudgetProjectState,
        ChangeOrder,
        Commitment,
        Company,
        DailyLog,
        Delivery,
        Document,
        Drawing,
        EquipmentEntry,
        ManpowerEntry,
        MeetingActionItem,
        MeetingMinute,
        PayAppProjectState,
        PermitInspectionItem,
        Photo,
        PotentialChangeOrder,
        Project,
        PunchItem,
        RFI,
        SafetyCertification,
        SafetyReport,
        ScheduleData,
        ScheduleTask,
        Submittal,
        User,
        WeeklyReport,
    )
    return {
        'Project': Project,
        'RFI': RFI,
        'ChangeOrder': ChangeOrder,
        'PotentialChangeOrder': PotentialChangeOrder,
        'Commitment': Commitment,
        'Submittal': Submittal,
        'PunchItem': PunchItem,
        'DailyLog': DailyLog,
        'ManpowerEntry': ManpowerEntry,
        'EquipmentEntry': EquipmentEntry,
        'WeeklyReport': WeeklyReport,
        'SafetyReport': SafetyReport,
        'SafetyCertification': SafetyCertification,
        'ScheduleData': ScheduleData,
        'ScheduleTask': ScheduleTask,
        'Delivery': Delivery,
        'PermitInspectionItem': PermitInspectionItem,
        'MeetingMinute': MeetingMinute,
        'MeetingActionItem': MeetingActionItem,
        'Photo': Photo,
        'Document': Document,
        'Drawing': Drawing,
        'BudgetProjectState': BudgetProjectState,
        'PayAppProjectState': PayAppProjectState,
        'Company': Company,
        'User': User,
        'AuditLog': AuditLog,
    }
